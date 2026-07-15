#!/usr/bin/env python3
"""One-time backfill of per-name skew history into ``skew_snapshots`` from an
externally-computed workbook (``skew_data.xlsx``).

WHY THIS EXISTS
    ``scheduler/jobs/skew_snapshots.py`` builds each day's row from that day's
    ``oi_chain_eod`` delta surface, and ``_latest_chain()`` ignores ``as_of`` —
    so the job cannot rebuild history. This importer bypasses the surface step
    and writes rows straight from the workbook's already-computed 25Δ call/put
    IVs, matching the job's schema, units, and upsert key exactly. Once loaded,
    the daily job's own 63d/252d percentiles read back this baseline, and
    ``get_skew_history`` / the Skew dashboard page light up immediately.

WHAT IT WRITES
    horizon_dte = 30 rows only (the workbook is a single 30d constant-maturity
    tenor; the daily job fills 60/90/180/365 forward). Columns populated:
        atm_iv, rr_25d (= vol_25p - vol_25c, decimal), bf_25d (= mean(wings) - atm),
        rr_25d_pctile_63d / _252d, bf_25d_pctile_252d  (0..1, min_history 20),
        shift_slide_label, label.
    Left NULL (not derivable from the workbook): rr_10d, bf_10d,
        front_back_rr_slope, vix_beta_60d, rr_25d_abnormal — the daily job fills
        these going forward.

RUN (laptop or NAS, repo venv active, from the repo root)
    python scripts/skew_backfill.py --file skew_data.xlsx --dry-run     # preview, no writes
    python scripts/skew_backfill.py --file skew_data.xlsx --through 2026-07-09
    # --through lets the 16:55 daily job own 2026-07-10 onward (avoids partial rows)

Idempotent: ON CONFLICT (symbol, ts, horizon_dte) DO UPDATE — safe to re-run.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert

from trading_intel.config import get_settings
from trading_intel.memory.db import make_session_factory
from trading_intel.memory.models import SkewSnapshot
from trading_intel.vol.skew import compose_label, shift_vs_slide, skew_percentile

HORIZON = 30
RR_SANITY = 0.25  # |vol_25p - vol_25c| (decimal) above this = degenerate fit -> drop
IV_SANITY = 3.0  # a single-wing IV above 300% = degenerate fit -> drop

# Column positions in each ticker sheet (row 1 = header).
COL = {"date": 0, "vol_25c": 3, "vol_25p": 4, "flags": 8, "excluded": 9, "atm": 11}

_UPDATE_COLS = (
    "atm_iv",
    "rr_10d",
    "rr_25d",
    "bf_10d",
    "bf_25d",
    "rr_25d_pctile_63d",
    "rr_25d_pctile_252d",
    "bf_25d_pctile_252d",
    "front_back_rr_slope",
    "vix_beta_60d",
    "rr_25d_abnormal",
    "shift_slide_label",
    "label",
)


def _parse_date(x: object) -> date | None:
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    if isinstance(x, str) and x.strip():
        return datetime.strptime(x[:10], "%Y-%m-%d").date()
    return None


def _is_excluded(v: object) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() not in ("", "no", "false", "0")


def clean_rows(ws: Worksheet) -> list[dict]:
    """Cleaned {ts, atm, rr, bf} for one ticker sheet, oldest-first, deduped by date."""
    out: dict[date, dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = _parse_date(r[COL["date"]])
        if d is None:
            continue
        v25c, v25p, atm = r[COL["vol_25c"]], r[COL["vol_25p"]], r[COL["atm"]]
        if not all(isinstance(v, (int, float)) for v in (v25c, v25p, atm)):
            continue  # need both wings + ATM (drops META/AVGO/DELL partial 7/10)
        if _is_excluded(r[COL["excluded"]]):
            continue  # respect the workbook's own sanity exclusions
        if v25c <= 0 or v25p <= 0 or atm <= 0:
            continue
        if v25c > IV_SANITY or v25p > IV_SANITY:
            continue  # blown-up wing
        rr = v25p - v25c
        if not (-RR_SANITY <= rr <= RR_SANITY):
            continue  # degenerate fit (e.g. the QQQ *_skew_bounds blow-ups)
        out[d] = {
            "ts": d,
            "atm": float(atm),
            "rr": float(rr),
            "bf": (float(v25c) + float(v25p)) / 2.0 - float(atm),
        }
    return [out[k] for k in sorted(out)]


def assemble(symbol: str, rows: list[dict]) -> list[dict]:
    """Build skew_snapshots row-dicts, computing percentiles/labels in date order."""
    records: list[dict] = []
    rr_hist: list[float] = []
    bf_hist: list[float] = []
    prev: dict | None = None
    for x in rows:
        pct63 = skew_percentile(rr_hist[-63:], x["rr"]) if rr_hist else None
        pct252 = skew_percentile(rr_hist[-252:], x["rr"]) if rr_hist else None
        bfp252 = skew_percentile(bf_hist[-252:], x["bf"]) if bf_hist else None
        if prev is not None:
            slide = shift_vs_slide(
                d_atm_iv_pts=(x["atm"] - prev["atm"]) * 100.0,
                d_rr_pts=(x["rr"] - prev["rr"]) * 100.0,
            )
        else:
            slide = None
        records.append(
            {
                "symbol": symbol,
                "ts": x["ts"],
                "horizon_dte": HORIZON,
                "atm_iv": x["atm"],
                "rr_10d": None,
                "rr_25d": x["rr"],
                "bf_10d": None,
                "bf_25d": x["bf"],
                "rr_25d_pctile_63d": pct63,
                "rr_25d_pctile_252d": pct252,
                "bf_25d_pctile_252d": bfp252,
                "front_back_rr_slope": None,
                "vix_beta_60d": None,
                "rr_25d_abnormal": None,
                "shift_slide_label": slide,
                "label": compose_label(rr_pts=x["rr"] * 100.0, pctile_long=pct252),
            }
        )
        rr_hist.append(x["rr"])
        bf_hist.append(x["bf"])
        prev = x
    return records


def main() -> None:
    ap = argparse.ArgumentParser(description="Backfill skew_snapshots from skew_data.xlsx")
    ap.add_argument("--file", required=True, help="path to skew_data.xlsx")
    ap.add_argument("--through", default=None, help="only import ts <= YYYY-MM-DD")
    ap.add_argument("--dry-run", action="store_true", help="preview counts, no DB writes")
    a = ap.parse_args()
    through = _parse_date(a.through) if a.through else None

    wb = openpyxl.load_workbook(a.file, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.lower() != "summary"]

    settings = get_settings()
    session_factory = make_session_factory(settings)
    with session_factory() as session:
        valid = set(session.execute(text("SELECT symbol FROM tickers")).scalars().all())
        exclude = settings.chain_exclude_roots  # index ETFs -> iv_tenor_snapshots

        records: list[dict] = []
        per_sym: dict[str, int] = {}
        skipped_fk: list[str] = []
        skipped_idx: list[str] = []
        for s in sheets:
            if s.upper() in exclude:
                skipped_idx.append(s)
                continue
            if s not in valid:
                skipped_fk.append(s)
                continue
            rows = clean_rows(wb[s])
            if through:
                rows = [r for r in rows if r["ts"] <= through]
            recs = assemble(s, rows)
            per_sym[s] = len(recs)
            records.extend(recs)

        print(
            f"importable symbols: {len(per_sym)} | rows: {len(records)}"
            + (f" | through {through}" if through else "")
        )
        for s in sorted(per_sym):
            print(f"  {s:6} {per_sym[s]:>4} rows")
        if skipped_idx:
            print(
                "skipped (index ETFs, CHAIN_EXCLUDE_ROOTS -> iv_tenor_snapshots): "
                + ", ".join(sorted(skipped_idx))
            )
        if skipped_fk:
            print("skipped (not in tickers table): " + ", ".join(sorted(skipped_fk)))

        if a.dry_run:
            print("DRY RUN — no rows written.")
            return

        for i in range(0, len(records), 1000):
            chunk = records[i : i + 1000]
            stmt = pg_insert(SkewSnapshot).values(chunk)
            stmt = stmt.on_conflict_do_update(
                index_elements=["symbol", "ts", "horizon_dte"],
                set_={c: stmt.excluded[c] for c in _UPDATE_COLS},
            )
            session.execute(stmt)
        session.commit()

        latest = session.execute(text("SELECT max(ts) FROM skew_snapshots")).scalar()
        names = session.execute(
            text("SELECT count(DISTINCT symbol) FROM skew_snapshots WHERE horizon_dte=30")
        ).scalar()
        print(
            f"UPSERTED {len(records)} rows. skew_snapshots latest ts={latest}, "
            f"names@30d={names}."
        )


if __name__ == "__main__":
    main()
